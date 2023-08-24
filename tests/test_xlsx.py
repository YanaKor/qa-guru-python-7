from openpyxl import load_workbook
import os

# TODO оформить в тест, добавить ассерты и использовать универсальный путь

xlsx_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'resources', 'file_example_XLSX_50.xlsx'))


def test_xlsx_file():
    workbook = load_workbook(xlsx_path)
    sheet = workbook.active

    assert sheet.cell(row=4, column=3).value == 'Gent'


def test_create_new_sheet():
    workbook = load_workbook(xlsx_path)
    workbook.create_sheet('New_sheet')

    assert workbook.sheetnames == ['Sheet1', 'New_sheet']
